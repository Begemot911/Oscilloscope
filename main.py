import matplotlib

matplotlib.use('TkAgg')
from matplotlib.backend_bases import MouseEvent
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import serial
from serial.tools import list_ports
import numpy as np
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from collections import deque
import time
import pandas as pd
import queue
import threading
import struct
from crc import Calculator, Crc8
import math
import neurokit2 as nk


class DemoECGGenerator:
    """Генератор демо-режима с реалистичными сигналами ЭКГ для 10 отведений"""

    def __init__(self, data_queue, sample_rate=250):
        self.data_queue = data_queue
        self.sample_rate = sample_rate
        self.is_running = False
        self.thread = None
        self.start_time = time.time()
        self.ecg_templates = self._create_ecg_templates()

    def _create_ecg_templates(self):
        """Создает шаблоны ЭКГ для разных отведений"""
        templates = {}

        # Базовый сердечный цикл (1 секунда = 60 уд/мин)
        t = np.linspace(0, 1.0, self.sample_rate)

        # Отведение I - стандартная ЭКГ
        templates[0] = self._generate_lead_i(t)
        # Отведение II - стандартная ЭКГ
        templates[1] = self._generate_lead_ii(t)
        # Отведение III - стандартная ЭКГ
        templates[2] = self._generate_lead_iii(t)
        # aVR - инвертированная
        templates[3] = self._generate_lead_avr(t)
        # aVL - малая амплитуда
        templates[4] = self._generate_lead_avl(t)
        # aVF - вертикальная ось
        templates[5] = self._generate_lead_avf(t)
        # V1 - грудное отведение
        templates[6] = self._generate_lead_v1(t)
        # V2 - грудное отведение
        templates[7] = self._generate_lead_v2(t)
        # V3 - грудное отведение
        templates[8] = self._generate_lead_v3(t)
        # V4 - грудное отведение
        templates[9] = self._generate_lead_v4(t)

        return templates

    def _generate_lead_i(self, t):
        """Генерирует сигнал для отведения I"""
        ecg = np.zeros_like(t)
        for i, time_val in enumerate(t):
            if 0.1 <= time_val < 0.15:
                ecg[i] = 0.1 * math.sin(2 * math.pi * (time_val - 0.1) / 0.05)
            elif 0.15 <= time_val < 0.25:
                ecg[i] = 0
            elif 0.25 <= time_val < 0.35:
                ecg[i] = 0.8 * math.sin(2 * math.pi * (time_val - 0.25) / 0.1) ** 3
            elif 0.35 <= time_val < 0.45:
                ecg[i] = 0.1
            elif 0.45 <= time_val < 0.65:
                ecg[i] = 0.3 * math.sin(math.pi * (time_val - 0.45) / 0.2) ** 2
            else:
                ecg[i] = 0
        return ecg

    def _generate_lead_ii(self, t):
        """Генерирует сигнал для отведения II"""
        ecg = np.zeros_like(t)
        for i, time_val in enumerate(t):
            if 0.1 <= time_val < 0.15:
                ecg[i] = 0.15 * math.sin(2 * math.pi * (time_val - 0.1) / 0.05)
            elif 0.15 <= time_val < 0.25:
                ecg[i] = 0
            elif 0.25 <= time_val < 0.35:
                ecg[i] = 1.2 * math.sin(2 * math.pi * (time_val - 0.25) / 0.1) ** 3
            elif 0.35 <= time_val < 0.45:
                ecg[i] = 0.15
            elif 0.45 <= time_val < 0.65:
                ecg[i] = 0.4 * math.sin(math.pi * (time_val - 0.45) / 0.2) ** 2
            else:
                ecg[i] = 0
        return ecg

    def _generate_lead_iii(self, t):
        """Генерирует сигнал для отведения III"""
        ecg = np.zeros_like(t)
        for i, time_val in enumerate(t):
            if 0.1 <= time_val < 0.15:
                ecg[i] = 0.12 * math.sin(2 * math.pi * (time_val - 0.1) / 0.05)
            elif 0.15 <= time_val < 0.25:
                ecg[i] = 0
            elif 0.25 <= time_val < 0.35:
                ecg[i] = 0.9 * math.sin(2 * math.pi * (time_val - 0.25) / 0.1) ** 3
            elif 0.35 <= time_val < 0.45:
                ecg[i] = 0.12
            elif 0.45 <= time_val < 0.65:
                ecg[i] = 0.35 * math.sin(math.pi * (time_val - 0.45) / 0.2) ** 2
            else:
                ecg[i] = 0
        return ecg

    def _generate_lead_avr(self, t):
        """Генерирует сигнал для отведения aVR"""
        ecg = np.zeros_like(t)
        for i, time_val in enumerate(t):
            if 0.1 <= time_val < 0.15:
                ecg[i] = -0.08 * math.sin(2 * math.pi * (time_val - 0.1) / 0.05)
            elif 0.15 <= time_val < 0.25:
                ecg[i] = 0
            elif 0.25 <= time_val < 0.35:
                ecg[i] = -0.6 * math.sin(2 * math.pi * (time_val - 0.25) / 0.1) ** 3
            elif 0.35 <= time_val < 0.45:
                ecg[i] = -0.08
            elif 0.45 <= time_val < 0.65:
                ecg[i] = -0.25 * math.sin(math.pi * (time_val - 0.45) / 0.2) ** 2
            else:
                ecg[i] = 0
        return ecg

    def _generate_lead_avl(self, t):
        """Генерирует сигнал для отведения aVL"""
        ecg = np.zeros_like(t)
        for i, time_val in enumerate(t):
            if 0.1 <= time_val < 0.15:
                ecg[i] = 0.05 * math.sin(2 * math.pi * (time_val - 0.1) / 0.05)
            elif 0.15 <= time_val < 0.25:
                ecg[i] = 0
            elif 0.25 <= time_val < 0.35:
                ecg[i] = 0.4 * math.sin(2 * math.pi * (time_val - 0.25) / 0.1) ** 3
            elif 0.35 <= time_val < 0.45:
                ecg[i] = 0.05
            elif 0.45 <= time_val < 0.65:
                ecg[i] = 0.15 * math.sin(math.pi * (time_val - 0.45) / 0.2) ** 2
            else:
                ecg[i] = 0
        return ecg

    def _generate_lead_avf(self, t):
        """Генерирует сигнал для отведения aVF"""
        ecg = np.zeros_like(t)
        for i, time_val in enumerate(t):
            if 0.1 <= time_val < 0.15:
                ecg[i] = 0.13 * math.sin(2 * math.pi * (time_val - 0.1) / 0.05)
            elif 0.15 <= time_val < 0.25:
                ecg[i] = 0
            elif 0.25 <= time_val < 0.35:
                ecg[i] = 1.0 * math.sin(2 * math.pi * (time_val - 0.25) / 0.1) ** 3
            elif 0.35 <= time_val < 0.45:
                ecg[i] = 0.13
            elif 0.45 <= time_val < 0.65:
                ecg[i] = 0.38 * math.sin(math.pi * (time_val - 0.45) / 0.2) ** 2
            else:
                ecg[i] = 0
        return ecg

    def _generate_lead_v1(self, t):
        """Генерирует сигнал для грудного отведения V1"""
        ecg = np.zeros_like(t)
        for i, time_val in enumerate(t):
            if 0.1 <= time_val < 0.15:
                ecg[i] = 0.08 * math.sin(2 * math.pi * (time_val - 0.1) / 0.05)
            elif 0.15 <= time_val < 0.25:
                ecg[i] = 0
            elif 0.25 <= time_val < 0.35:
                if time_val < 0.28:
                    ecg[i] = 0.3 * math.sin(2 * math.pi * (time_val - 0.25) / 0.03) ** 2
                elif time_val < 0.32:
                    ecg[i] = -0.2
                else:
                    ecg[i] = 0.4 * math.sin(2 * math.pi * (time_val - 0.32) / 0.03) ** 2
            elif 0.35 <= time_val < 0.45:
                ecg[i] = 0.05
            elif 0.45 <= time_val < 0.65:
                ecg[i] = 0.2 * math.sin(math.pi * (time_val - 0.45) / 0.2) ** 2
            else:
                ecg[i] = 0
        return ecg

    def _generate_lead_v2(self, t):
        """Генерирует сигнал для грудного отведения V2"""
        ecg = np.zeros_like(t)
        for i, time_val in enumerate(t):
            if 0.1 <= time_val < 0.15:
                ecg[i] = 0.1 * math.sin(2 * math.pi * (time_val - 0.1) / 0.05)
            elif 0.15 <= time_val < 0.25:
                ecg[i] = 0
            elif 0.25 <= time_val < 0.35:
                ecg[i] = 0.9 * math.sin(2 * math.pi * (time_val - 0.25) / 0.1) ** 3
            elif 0.35 <= time_val < 0.45:
                ecg[i] = 0.1
            elif 0.45 <= time_val < 0.65:
                ecg[i] = 0.35 * math.sin(math.pi * (time_val - 0.45) / 0.2) ** 2
            else:
                ecg[i] = 0
        return ecg

    def _generate_lead_v3(self, t):
        """Генерирует сигнал для грудного отведения V3"""
        ecg = np.zeros_like(t)
        for i, time_val in enumerate(t):
            if 0.1 <= time_val < 0.15:
                ecg[i] = 0.09 * math.sin(2 * math.pi * (time_val - 0.1) / 0.05)
            elif 0.15 <= time_val < 0.25:
                ecg[i] = 0
            elif 0.25 <= time_val < 0.35:
                ecg[i] = 0.7 * math.sin(2 * math.pi * (time_val - 0.25) / 0.1) ** 3
            elif 0.35 <= time_val < 0.45:
                ecg[i] = 0.08
            elif 0.45 <= time_val < 0.65:
                ecg[i] = 0.25 * math.sin(math.pi * (time_val - 0.45) / 0.2) ** 2
            else:
                ecg[i] = 0
        return ecg

    def _generate_lead_v4(self, t):
        """Генерирует сигнал для грудного отведения V4"""
        ecg = np.zeros_like(t)
        for i, time_val in enumerate(t):
            if 0.1 <= time_val < 0.15:
                ecg[i] = 0.11 * math.sin(2 * math.pi * (time_val - 0.1) / 0.05)
            elif 0.15 <= time_val < 0.25:
                ecg[i] = 0
            elif 0.25 <= time_val < 0.35:
                ecg[i] = 0.6 * math.sin(2 * math.pi * (time_val - 0.25) / 0.1) ** 3
            elif 0.35 <= time_val < 0.45:
                ecg[i] = 0.09
            elif 0.45 <= time_val < 0.65:
                ecg[i] = 0.22 * math.sin(math.pi * (time_val - 0.45) / 0.2) ** 2
            else:
                ecg[i] = 0
        return ecg

    def _add_respiration_artifact(self, value, timestamp, channel):
        """Добавляет артефакты дыхания"""
        respiration_freq = 0.2
        return value + 0.05 * math.sin(2 * math.pi * respiration_freq * timestamp + channel * 0.5)

    def _add_muscle_noise(self, value):
        """Добавляет мышечный шум"""
        return value + 0.02 * (np.random.random() - 0.5)

    def _add_baseline_wander(self, value, timestamp):
        """Добавляет дрейф изолинии"""
        wander_freq = 0.05
        return value + 0.03 * math.sin(2 * math.pi * wander_freq * timestamp)

    def start(self):
        """Запускает генерацию демо-данных"""
        if self.is_running:
            return False

        self.is_running = True
        self.start_time = time.time()
        self.thread = threading.Thread(target=self._generate_data, daemon=True)
        self.thread.start()
        print("Demo mode started - generating 10-channel ECG data")
        return True

    def _generate_data(self):
        """Генерирует демо-данные в отдельном потоке"""
        sample_interval = 1.0 / self.sample_rate
        last_sample_time = self.start_time

        while self.is_running:
            current_time = time.time()
            elapsed = current_time - last_sample_time

            if elapsed >= sample_interval:
                timestamp = current_time - self.start_time

                for channel in range(10):
                    cycle_pos = timestamp % 1.0
                    template = self.ecg_templates[channel]
                    idx = int(cycle_pos * self.sample_rate) % len(template)
                    value = template[idx]

                    value = self._add_respiration_artifact(value, timestamp, channel)
                    value = self._add_muscle_noise(value)
                    value = self._add_baseline_wander(value, timestamp)

                    self.data_queue.put({
                        'port': 'DEMO',
                        'channel': channel + 1,
                        'timestamp': timestamp,
                        'value': value
                    })

                last_sample_time = current_time
            else:
                time.sleep(0.001)

    def stop(self):
        """Останавливает генерацию демо-данных"""
        self.is_running = False
        if self.thread and self.thread.is_alive():
            self.thread.join(timeout=1.0)
        print("Demo mode stopped")


class ECGRealTimeAnalyzer:
    """Анализатор ЭКГ в реальном времени"""

    def __init__(self, channel_data, channel_idx, sampling_rate=250):
        self.channel_data = channel_data
        self.channel_idx = channel_idx
        self.sampling_rate = sampling_rate
        self.analysis_window_seconds = 10

    def analyze(self):
        """Анализирует данные выбранного канала"""
        ch = self.channel_data[self.channel_idx]

        if len(ch['timestamps']) < 100:
            return {
                'error': f'Недостаточно данных для анализа канала {self.channel_idx + 1}. Нужно минимум 100 точек.'
            }

        if len(ch['timestamps']) > 0:
            last_time = ch['timestamps'][-1]
            cutoff_time = last_time - self.analysis_window_seconds

            timestamps = np.array(ch['timestamps'])
            indices = np.where(timestamps >= cutoff_time)[0]

            if len(indices) < 100:
                indices = range(len(ch['timestamps']))

            time_data = np.array([ch['timestamps'][i] for i in indices])
            ecg_signal = np.array([ch['filtered_data'][i] for i in indices])
        else:
            return {'error': 'Нет данных для анализа'}

        if len(time_data) > 1:
            diffs = np.diff(time_data)
            if len(diffs) > 0 and np.mean(diffs) > 0:
                self.sampling_rate = 1 / np.mean(diffs)
            else:
                self.sampling_rate = 250

        try:
            ecg_clean = nk.ecg_clean(ecg_signal, sampling_rate=self.sampling_rate)
            _, rpeaks = nk.ecg_peaks(ecg_clean, sampling_rate=self.sampling_rate)
            peaks = rpeaks['ECG_R_Peaks']

            if len(peaks) < 2:
                return {
                    'error': f'Обнаружено только {len(peaks)} R-пиков. Нужно минимум 2 для анализа.',
                    'peaks_count': len(peaks)
                }

            rr_intervals = np.diff(peaks) / self.sampling_rate
            heart_rate = 60 / np.mean(rr_intervals)
            hr_min = 60 / np.max(rr_intervals) if len(rr_intervals) > 0 else heart_rate
            hr_max = 60 / np.min(rr_intervals) if len(rr_intervals) > 0 else heart_rate
            sdnn = np.std(rr_intervals) * 1000
            rmssd = np.sqrt(np.mean(np.diff(rr_intervals) ** 2)) * 1000
            arrhythmia_detected = sdnn > 50
            abnormalities = self._detect_abnormalities(ecg_clean, peaks, self.sampling_rate)

            return {
                'success': True,
                'channel': self.channel_idx + 1,
                'duration': time_data[-1] - time_data[0] if len(time_data) > 1 else 0,
                'sampling_rate': self.sampling_rate,
                'points_count': len(ecg_signal),
                'peaks_count': len(peaks),
                'heart_rate': heart_rate,
                'hr_min': hr_min,
                'hr_max': hr_max,
                'sdnn': sdnn,
                'rmssd': rmssd,
                'arrhythmia_detected': arrhythmia_detected,
                'abnormalities': abnormalities
            }

        except Exception as e:
            return {'error': f'Ошибка анализа: {str(e)}'}

    def _detect_abnormalities(self, ecg_signal, peaks, sampling_rate):
        """Детектирует возможные патологии"""
        abnormalities = []

        if len(peaks) < 3:
            return abnormalities

        qrs_durations = []
        for peak in peaks:
            if peak > 5 and peak < len(ecg_signal) - 5:
                left = peak
                while left > peak - 30 and left > 0 and ecg_signal[left] > 0.1 * max(ecg_signal[peak - 5:peak + 5]):
                    left -= 1
                right = peak
                while right < peak + 30 and right < len(ecg_signal) - 1 and ecg_signal[right] > 0.1 * max(
                        ecg_signal[peak - 5:peak + 5]):
                    right += 1
                qrs_duration = (right - left) / sampling_rate * 1000
                qrs_durations.append(qrs_duration)

        if qrs_durations:
            avg_qrs = np.mean(qrs_durations)
            if avg_qrs > 120:
                abnormalities.append(f"Удлинение QRS ({avg_qrs:.0f} мс > 120 мс)")

        if len(peaks) > 1:
            rr_intervals = np.diff(peaks) / sampling_rate
            heart_rate = 60 / np.mean(rr_intervals)

            if heart_rate > 100:
                abnormalities.append(f"Тахикардия (ЧСС = {heart_rate:.0f} уд/мин)")
            elif heart_rate < 60:
                abnormalities.append(f"Брадикардия (ЧСС = {heart_rate:.0f} уд/мин)")

        return abnormalities


class SerialReader:
    def __init__(self, port, baudrate, data_queue, crc_calculator):
        self.port = port
        self.baudrate = baudrate
        self.data_queue = data_queue
        self.crc_calculator = crc_calculator
        self.ser = None
        self.is_running = False
        self.thread = None
        self.buffer = bytearray()
        self.lock = threading.Lock()

    def start(self):
        if self.is_running:
            return False

        try:
            print(f"Connecting to {self.port} at {self.baudrate} baud...")
            self.ser = serial.Serial(
                self.port,
                self.baudrate,
                timeout=0.01
            )
            self.ser.reset_input_buffer()
            self.is_running = True
            self.thread = threading.Thread(target=self.read_from_port, daemon=True)
            self.thread.start()
            print(f"Successfully connected to {self.port}")
            return True
        except Exception as e:
            print(f"Failed to connect to {self.port}: {str(e)}")
            return False

    def read_from_port(self):
        PACKET_HEADER = 0xAA
        PACKET_SIZE = 11

        print(f"Starting data processing on {self.port}")

        while self.is_running and self.ser and self.ser.is_open:
            try:
                bytes_to_read = self.ser.in_waiting
                if bytes_to_read == 0:
                    time.sleep(0.01)
                    continue

                raw_data = self.ser.read(bytes_to_read)
                self.buffer.extend(raw_data)

                while len(self.buffer) >= PACKET_SIZE:
                    header_pos = -1
                    for i in range(len(self.buffer) - PACKET_SIZE + 1):
                        if self.buffer[i] == PACKET_HEADER:
                            header_pos = i
                            break

                    if header_pos == -1:
                        self.buffer.clear()
                        break

                    if header_pos > 0:
                        print(f"{self.port}: Discarded {header_pos} bytes before header")
                        del self.buffer[:header_pos]
                        continue

                    packet = bytes(self.buffer[:PACKET_SIZE])
                    computed_crc = self.crc_calculator.checksum(packet[:10])

                    if computed_crc != packet[10]:
                        print(f"{self.port}: CRC error, skipping packet")
                        del self.buffer[:1]
                        continue

                    channel = packet[1]
                    timestamp = struct.unpack('<f', packet[2:6])[0]
                    value = struct.unpack('<f', packet[6:10])[0]

                    self.data_queue.put({
                        'port': self.port,
                        'channel': channel,
                        'timestamp': timestamp,
                        'value': value
                    })

                    del self.buffer[:PACKET_SIZE]

            except serial.SerialException as e:
                print(f"{self.port}: Serial error - {str(e)}")
                break
            except Exception as e:
                print(f"{self.port}: Processing error - {str(e)}")
                time.sleep(0.1)

        print(f"Stopped reading from {self.port}")

    def send_command(self, command):
        if self.ser and self.ser.is_open:
            try:
                self.ser.write(command.encode('utf-8'))
                self.ser.flush()
                print(f"Sent command '{command}' to {self.port}")
                return True
            except Exception as e:
                print(f"Failed to send command to {self.port}: {str(e)}")
        return False

    def stop(self):
        with self.lock:
            self.is_running = False

        if self.thread and self.thread.is_alive():
            self.thread.join(timeout=1.0)

        if self.ser and self.ser.is_open:
            try:
                self.ser.close()
                print(f"Port {self.port} closed successfully")
            except Exception as e:
                print(f"Error closing port {self.port}: {str(e)}")

        self.buffer.clear()
        print(f"Port {self.port} fully stopped")


class Oscilloscope:
    def __init__(self, master):
        self.crc_calculator = Calculator(Crc8.CCITT)
        self.after_id = None
        self.master = master
        self.serial_readers = []
        self.demo_generator = None
        self.is_running = False
        self.paused = False
        self.buffer_size = 5000
        self.total_points = 0
        self.num_channels = 10
        self.channels = []
        self.demo_mode = False

        colors = ['blue', 'green', 'red', 'purple', 'darkorange',
                  'navy', 'magenta', 'cyan', 'brown', 'pink']

        for i in range(self.num_channels):
            self.channels.append({
                'color': colors[i % len(colors)],
                'raw_data': deque(maxlen=self.buffer_size),
                'filtered_data': deque(maxlen=self.buffer_size),
                'timestamps': deque(maxlen=self.buffer_size),
                'lpf_states': None,
                'hpf_states': None,
                'bpf_sections': None
            })

        self.start_time = None
        self.data_queue = queue.Queue()
        self.filter_params = {
            'type': 'None',
            'cutoff': 50.0
        }
        self.buffer = bytearray()

        self.setup_gui()
        self.setup_plots()
        self.refresh_ports()

        self.master.bind('+', self.increase_scale)
        self.master.bind('-', self.decrease_scale)

        master.protocol("WM_DELETE_WINDOW", self.on_close)

    def on_close(self):
        print("Closing application...")
        self.stop()
        try:
            self.data_queue.queue.clear()
            for ch in self.channels:
                ch['raw_data'].clear()
                ch['filtered_data'].clear()
                ch['timestamps'].clear()
        except:
            pass
        self.master.destroy()
        print("Application closed")

    def setup_gui(self):
        main_panel = ttk.PanedWindow(self.master, orient=tk.HORIZONTAL)
        main_panel.pack(fill=tk.BOTH, expand=True)

        left_info_panel = ttk.Frame(main_panel, width=250)
        main_panel.add(left_info_panel)

        right_panel = ttk.Frame(main_panel)
        main_panel.add(right_panel)

        patient_frame = ttk.LabelFrame(left_info_panel, text="Информация о пациенте")
        patient_frame.pack(padx=10, pady=10, fill=tk.X)

        ttk.Label(patient_frame, text="Фамилия:").grid(row=0, column=0, sticky=tk.W)
        self.surname_entry = ttk.Entry(patient_frame)
        self.surname_entry.grid(row=0, column=1, sticky=tk.EW, padx=5, pady=2)

        ttk.Label(patient_frame, text="Имя:").grid(row=1, column=0, sticky=tk.W)
        self.name_entry = ttk.Entry(patient_frame)
        self.name_entry.grid(row=1, column=1, sticky=tk.EW, padx=5, pady=2)

        ttk.Label(patient_frame, text="Отчество:").grid(row=2, column=0, sticky=tk.W)
        self.patronymic_entry = ttk.Entry(patient_frame)
        self.patronymic_entry.grid(row=2, column=1, sticky=tk.EW, padx=5, pady=2)

        ttk.Label(patient_frame, text="Пол:").grid(row=3, column=0, sticky=tk.W)
        self.gender_combo = ttk.Combobox(patient_frame, values=["М", "Ж"], width=3)
        self.gender_combo.grid(row=3, column=1, sticky=tk.W, padx=5, pady=2)

        ttk.Label(patient_frame, text="Год рождения:").grid(row=4, column=0, sticky=tk.W)
        self.birth_year_entry = ttk.Entry(patient_frame, width=8)
        self.birth_year_entry.grid(row=4, column=1, sticky=tk.W, padx=5, pady=2)

        ttk.Label(patient_frame, text="Диагноз:").grid(row=5, column=0, sticky=tk.W)
        self.diagnosis_entry = ttk.Entry(patient_frame)
        self.diagnosis_entry.grid(row=5, column=1, sticky=tk.EW, padx=5, pady=2)

        # Выбор канала для анализа
        channel_select_frame = ttk.Frame(patient_frame)
        channel_select_frame.grid(row=6, column=0, columnspan=2, sticky=tk.EW, pady=5)

        ttk.Label(channel_select_frame, text="Канал для анализа:").pack(side=tk.LEFT, padx=5)
        self.channel_selector = ttk.Combobox(channel_select_frame,
                                             values=[str(i) for i in range(1, 11)],
                                             width=5, state="readonly")
        self.channel_selector.set("1")
        self.channel_selector.pack(side=tk.LEFT, padx=5)

        # Кнопка анализа ЭКГ
        self.analyze_btn = ttk.Button(patient_frame, text="Анализ ЭКГ", command=self.analyze_selected_channel)
        self.analyze_btn.grid(row=7, column=0, columnspan=2, sticky=tk.EW, pady=5)

        results_frame = ttk.LabelFrame(left_info_panel, text="Результаты анализа")
        results_frame.pack(padx=10, pady=10, fill=tk.BOTH, expand=True)

        self.result_text = tk.Text(results_frame, height=10, wrap=tk.WORD, width=40)
        scrollbar = ttk.Scrollbar(results_frame, command=self.result_text.yview)
        self.result_text.configure(yscrollcommand=scrollbar.set)

        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.result_text.pack(fill=tk.BOTH, expand=False)

        self.notebook = ttk.Notebook(right_panel)
        self.notebook.pack(fill=tk.BOTH, expand=True)

        time_frame = ttk.Frame(self.notebook)
        self.notebook.add(time_frame, text='Oscilloscope')

        control_frame = ttk.Frame(time_frame)
        control_frame.pack(side=tk.TOP, fill=tk.X, padx=5, pady=5)

        left_panel = ttk.Frame(control_frame)
        left_panel.pack(side=tk.LEFT, fill=tk.X, expand=True)

        right_panel_controls = ttk.Frame(control_frame)
        right_panel_controls.pack(side=tk.RIGHT)

        ports_frame = ttk.Frame(left_panel)
        ports_frame.pack(side=tk.TOP, fill=tk.X)

        # Создаем фреймы для каждого порта
        port_frames = []
        for i in range(6):
            port_frame = ttk.Frame(ports_frame)
            port_frame.pack(side=tk.LEFT, padx=2)
            ttk.Label(port_frame, text=f"Port {i + 1}:").pack(side=tk.LEFT)
            combo = ttk.Combobox(port_frame, width=12)
            combo.pack(side=tk.LEFT)
            port_frames.append(combo)

        # Сохраняем ссылки на комбобоксы
        self.port1_combo, self.port2_combo, self.port3_combo, self.port4_combo, self.port5_combo, self.port6_combo = port_frames

        settings_frame = ttk.Frame(left_panel)
        settings_frame.pack(side=tk.TOP, fill=tk.X)

        ttk.Label(settings_frame, text="Baud:").pack(side=tk.LEFT)
        self.baud_combo = ttk.Combobox(settings_frame,
                                       values=[9600, 19200, 38400, 57600, 115200, 250000, 500000, 1000000],
                                       width=10)
        self.baud_combo.current(6)
        self.baud_combo.pack(side=tk.LEFT)

        # Кнопка демо-режима с новыми надписями
        self.demo_btn = ttk.Button(settings_frame, text="▶ Start Demo", command=self.toggle_demo)
        self.demo_btn.pack(side=tk.LEFT, padx=5)

        self.connect_btn = ttk.Button(settings_frame, text="Connect", command=self.toggle_connection)
        self.connect_btn.pack(side=tk.LEFT, padx=5)

        ttk.Label(settings_frame, text="Points:").pack(side=tk.LEFT)
        self.points_entry = ttk.Entry(settings_frame, width=8)
        self.points_entry.insert(0, "1000")
        self.points_entry.pack(side=tk.LEFT)

        ttk.Label(settings_frame, text="Filter:").pack(side=tk.LEFT)
        self.filter_combo = ttk.Combobox(settings_frame, values=['None', 'LPF', 'HPF', 'BPF'], width=6)
        self.filter_combo.current(0)
        self.filter_combo.pack(side=tk.LEFT)
        self.filter_combo.bind("<<ComboboxSelected>>", self.update_filter_ui)

        ttk.Label(settings_frame, text="Order:").pack(side=tk.LEFT, padx=(10, 0))
        self.filter_order_combo = ttk.Combobox(settings_frame, values=[1, 2, 4], width=3)
        self.filter_order_combo.current(0)
        self.filter_order_combo.pack(side=tk.LEFT)

        self.std_cutoff_frame = ttk.Frame(settings_frame)
        self.std_cutoff_label = ttk.Label(self.std_cutoff_frame, text="Cutoff (Hz):")
        self.std_cutoff_label.pack(side=tk.LEFT)
        self.std_cutoff_entry = ttk.Entry(self.std_cutoff_frame, width=8)
        self.std_cutoff_entry.insert(0, "50")
        self.std_cutoff_entry.pack(side=tk.LEFT)
        self.std_cutoff_frame.pack(side=tk.LEFT)

        self.bpf_cutoff_frame = ttk.Frame(settings_frame)
        self.bpf_low_label = ttk.Label(self.bpf_cutoff_frame, text="Low:")
        self.bpf_low_label.pack(side=tk.LEFT)
        self.bpf_low_entry = ttk.Entry(self.bpf_cutoff_frame, width=6)
        self.bpf_low_entry.insert(0, "20")
        self.bpf_low_entry.pack(side=tk.LEFT)
        self.bpf_high_label = ttk.Label(self.bpf_cutoff_frame, text="High:")
        self.bpf_high_label.pack(side=tk.LEFT)
        self.bpf_high_entry = ttk.Entry(self.bpf_cutoff_frame, width=6)
        self.bpf_high_entry.insert(0, "100")
        self.bpf_high_entry.pack(side=tk.LEFT)
        self.bpf_cutoff_frame.pack_forget()

        self.export_btn = ttk.Button(right_panel_controls, text="📤 Export", command=self.export_data)
        self.export_btn.pack(side=tk.RIGHT, padx=5)
        self.export_btn.state(['disabled'])

        self.pause_btn = ttk.Button(right_panel_controls, text="⏸ Stop", command=self.toggle_pause)
        self.pause_btn.pack(side=tk.RIGHT, padx=5)
        self.pause_btn.state(['disabled'])

    def analyze_selected_channel(self):
        """Анализирует выбранный канал ЭКГ"""
        channel_num = self.channel_selector.get()
        if not channel_num:
            messagebox.showwarning("Выбор канала", "Пожалуйста, выберите номер канала для анализа")
            return

        channel_idx = int(channel_num) - 1

        if channel_idx < 0 or channel_idx >= self.num_channels:
            messagebox.showerror("Ошибка", f"Неверный номер канала. Доступны каналы 1-{self.num_channels}")
            return

        if len(self.channels[channel_idx]['timestamps']) < 100:
            messagebox.showwarning("Недостаточно данных",
                                   f"Для канала {channel_num} недостаточно данных для анализа. "
                                   f"Текущее количество точек: {len(self.channels[channel_idx]['timestamps'])}. "
                                   "Нужно минимум 100 точек.")
            return

        try:
            analyzer = ECGRealTimeAnalyzer(self.channels, channel_idx)
            result = analyzer.analyze()

            if 'error' in result:
                messagebox.showerror("Ошибка анализа", result['error'])
                return

            result_text = (
                f"=== Анализ ЭКГ (Канал {result['channel']}) ===\n\n"
                f"Параметры сигнала:\n"
                f"  • Длительность записи: {result['duration']:.1f} сек\n"
                f"  • Частота дискретизации: {result['sampling_rate']:.1f} Гц\n"
                f"  • Количество точек: {result['points_count']}\n"
                f"  • Обнаружено R-пиков: {result['peaks_count']}\n\n"
                f"Показатели сердечного ритма:\n"
                f"  • Средняя ЧСС: {result['heart_rate']:.1f} уд/мин\n"
                f"  • Минимальная ЧСС: {result['hr_min']:.1f} уд/мин\n"
                f"  • Максимальная ЧСС: {result['hr_max']:.1f} уд/мин\n"
                f"  • SDNN (вариабельность): {result['sdnn']:.1f} мс\n"
                f"  • RMSSD: {result['rmssd']:.1f} мс\n\n"
            )

            if result['arrhythmia_detected']:
                result_text += "⚠️ ВНИМАНИЕ: Обнаружена возможная аритмия (высокая вариабельность RR-интервалов)\n"

            if result['abnormalities']:
                result_text += "\n🔍 Обнаруженные отклонения:\n"
                for abnormality in result['abnormalities']:
                    result_text += f"  • {abnormality}\n"
            else:
                result_text += "\n✅ Значительных отклонений не обнаружено\n"

            self.result_text.delete(1.0, tk.END)
            self.result_text.insert(tk.END, result_text)

            self.show_analysis_plot(self.channels[channel_idx], channel_idx)

        except Exception as e:
            messagebox.showerror("Ошибка анализа", f"Не удалось выполнить анализ:\n{str(e)}")

    def show_analysis_plot(self, channel_data, channel_idx):
        """Показывает детальный график анализа выбранного канала"""
        plt.close('all')

        timestamps = np.array(channel_data['timestamps'])
        ecg_signal = np.array(channel_data['filtered_data'])

        if len(timestamps) > 0:
            last_time = timestamps[-1]
            cutoff_time = last_time - 10
            indices = np.where(timestamps >= cutoff_time)[0]

            if len(indices) < 100:
                indices = range(len(timestamps))

            time_data = timestamps[indices]
            ecg_data = ecg_signal[indices]
        else:
            messagebox.showwarning("Нет данных", "Недостаточно данных для построения графика")
            return

        if len(time_data) > 1:
            sampling_rate = 1 / np.mean(np.diff(time_data))
        else:
            sampling_rate = 250

        try:
            ecg_clean = nk.ecg_clean(ecg_data, sampling_rate=sampling_rate)
            _, rpeaks = nk.ecg_peaks(ecg_clean, sampling_rate=sampling_rate)
            peaks = rpeaks['ECG_R_Peaks']

            fig, (ax1, ax2, ax3) = plt.subplots(3, 1, figsize=(12, 10))

            ax1.plot(time_data, ecg_clean, 'b-', label='ЭКГ сигнал', linewidth=1)
            if len(peaks) > 0:
                peak_times = time_data[peaks]
                peak_values = ecg_clean[peaks]
                ax1.scatter(peak_times, peak_values, color='red', s=30, zorder=5, label='R-пики')

            ax1.set_title(f'Анализ ЭКГ (Канал {channel_idx + 1})')
            ax1.set_xlabel('Время (с)')
            ax1.set_ylabel('Амплитуда')
            ax1.legend()
            ax1.grid(True, alpha=0.3)

            if len(peaks) > 1:
                rr_intervals = np.diff(peaks) / sampling_rate
                rr_times = time_data[peaks[1:]]

                ax2.stem(rr_times, rr_intervals, basefmt=" ", markerfmt='ro', linefmt='r-')
                ax2.set_title('RR-интервалы')
                ax2.set_xlabel('Время (с)')
                ax2.set_ylabel('Интервал (с)')
                ax2.grid(True, alpha=0.3)

                mean_rr = np.mean(rr_intervals)
                ax2.axhline(y=mean_rr, color='g', linestyle='--', label=f'Средний: {mean_rr:.3f} с')
                ax2.legend()

            if len(peaks) > 1:
                heart_rates = 60 / rr_intervals
                hr_times = rr_times

                ax3.plot(hr_times, heart_rates, 'g-', linewidth=2)
                ax3.fill_between(hr_times, heart_rates, alpha=0.3)
                ax3.set_title('Частота сердечных сокращений')
                ax3.set_xlabel('Время (с)')
                ax3.set_ylabel('ЧСС (уд/мин)')
                ax3.grid(True, alpha=0.3)

                ax3.axhline(y=60, color='orange', linestyle='--', alpha=0.7, label='Норма (60-100)')
                ax3.axhline(y=100, color='orange', linestyle='--', alpha=0.7)
                ax3.legend()

            plt.tight_layout()
            plt.show()

        except Exception as e:
            messagebox.showerror("Ошибка", f"Не удалось построить график анализа:\n{str(e)}")

    def toggle_demo(self):
        """Переключает демо-режим"""
        if self.demo_mode:
            self.stop_demo()
        else:
            self.start_demo()

    def start_demo(self):
        """Запускает демо-режим"""
        if self.is_running:
            self.stop()

        self._reset_measurement()

        self.demo_generator = DemoECGGenerator(self.data_queue)

        if self.demo_generator.start():
            self.demo_mode = True
            self.is_running = True
            self.paused = False
            self.demo_btn.config(text="⏹ Stop Demo")
            self.connect_btn.state(['disabled'])
            self.pause_btn.state(['!disabled'])
            self.export_btn.state(['!disabled'])

            for i in range(10):
                self.channels[i]['port'] = 'DEMO'

            self.update_plot()

            print("Demo mode started - simulating 10-channel ECG")
        else:
            messagebox.showerror("Error", "Failed to start demo mode")

    def stop_demo(self):
        """Останавливает демо-режим"""
        if self.demo_generator:
            self.demo_generator.stop()
            self.demo_generator = None

        self.demo_mode = False
        self.is_running = False
        self.demo_btn.config(text="▶ Start Demo")
        self.connect_btn.state(['!disabled'])
        self.pause_btn.state(['disabled'])
        self.export_btn.state(['disabled'])
        print("Demo mode stopped")

    def setup_plots(self):
        time_frame = self.notebook.winfo_children()[0]
        self.fig_time = plt.figure(figsize=(12, 10), dpi=100)
        self.axes = []
        self.lines = []

        for i in range(self.num_channels):
            row = i % 5
            col = i // 5
            ax = self.fig_time.add_subplot(5, 2, i + 1)
            line, = ax.plot([], [], lw=1, color=self.channels[i]['color'])
            ax.set_ylabel(f'Ch {i + 1}')
            ax.grid(True)
            self.axes.append(ax)
            self.lines.append(line)

        self.annotations = []
        for i, ax in enumerate(self.axes):
            annotation = ax.annotate('',
                                     xy=(0, 0),
                                     xytext=(5, -15 if i < self.num_channels - 1 else 5),
                                     textcoords='offset points',
                                     bbox=dict(boxstyle="round", fc="w", alpha=0.8),
                                     arrowprops=dict(arrowstyle="->"))
            annotation.set_visible(False)
            self.annotations.append(annotation)

        self.fig_time.canvas.mpl_connect('motion_notify_event', self.on_mouse_move)
        self.fig_time.canvas.mpl_connect('figure_leave_event', self.on_leave_figure)

        for i in range(self.num_channels - 2, self.num_channels):
            self.axes[i].set_xlabel('Time (seconds)')

        self.canvas_time = FigureCanvasTkAgg(self.fig_time, master=time_frame)
        self.canvas_time.get_tk_widget().pack(fill=tk.BOTH, expand=True)
        self.fig_time.tight_layout()

    def on_mouse_move(self, event):
        if event.inaxes is None:
            for ann in self.annotations:
                ann.set_visible(False)
            self.canvas_time.draw_idle()
            return

        for i, ax in enumerate(self.axes):
            if ax == event.inaxes:
                line = self.lines[i]
                xdata = line.get_xdata()
                ydata = line.get_ydata()

                if len(xdata) == 0:
                    continue

                idx = np.abs(xdata - event.xdata).argmin()
                x = xdata[idx]
                y = ydata[idx]

                self.annotations[i].xy = (x, y)
                self.annotations[i].set_text(f'x={x:.2f}s, y={y:.2f}')
                self.annotations[i].set_visible(True)
            else:
                self.annotations[i].set_visible(False)

        self.canvas_time.draw_idle()

    def on_leave_figure(self, event):
        for ann in self.annotations:
            ann.set_visible(False)
        self.canvas_time.draw_idle()

    def update_filter_ui(self, event=None):
        filter_type = self.filter_combo.get()
        if filter_type == 'BPF':
            self.std_cutoff_frame.pack_forget()
            self.bpf_cutoff_frame.pack(side=tk.LEFT)
        else:
            self.bpf_cutoff_frame.pack_forget()
            self.std_cutoff_frame.pack(side=tk.LEFT)

    def refresh_ports(self):
        ports = [port.device for port in list_ports.comports()]
        port_values = ["None"] + ports
        for combo in [self.port1_combo, self.port2_combo, self.port3_combo,
                      self.port4_combo, self.port5_combo, self.port6_combo]:
            combo['values'] = port_values
            combo.current(0)
        self.port1_combo.current(1 if len(ports) > 0 else 0)
        self.port2_combo.current(2 if len(ports) > 1 else 0)

    def send_start_command(self):
        """Отправляет команду 'start' на все подключенные порты"""
        if not self.serial_readers:
            return False

        results = []
        for reader in self.serial_readers:
            if reader.is_running and reader.ser and reader.ser.is_open:
                results.append(reader.send_command("start"))

        return any(results)

    def send_stop_command(self):
        """Отправляет команду 'stop' на все подключенные порты"""
        if not self.serial_readers:
            return False

        results = []
        for reader in self.serial_readers:
            if reader.is_running and reader.ser and reader.ser.is_open:
                results.append(reader.send_command("stop"))

        return any(results)

    def toggle_connection(self):
        if self.is_running:
            self.stop()
        else:
            self.start()

    def toggle_pause(self):
        if self.paused:
            self._reset_measurement()
            if not self.demo_mode:
                for reader in self.serial_readers:
                    if reader.ser and reader.ser.is_open:
                        reader.ser.reset_input_buffer()
            self.paused = False
            self.pause_btn.config(text="⏸ Stop")
            if not self.demo_mode:
                self.send_start_command()
        else:
            if not self.demo_mode:
                self.send_stop_command()
            self.paused = True
            self.pause_time = time.time()
            self.pause_btn.config(text="▶ Start")

    def _reset_measurement(self):
        self.total_points = 0
        self.data_queue.queue.clear()

        for ch in self.channels:
            ch['raw_data'].clear()
            ch['filtered_data'].clear()
            ch['timestamps'].clear()
            ch['lpf_states'] = None
            ch['hpf_states'] = None
            ch['bpf_sections'] = None

        for line in self.lines:
            line.set_data([], [])
        self.canvas_time.draw()

    def start(self):
        try:
            self.stop()

            self._reset_measurement()

            ports = [
                self.port1_combo.get(),
                self.port2_combo.get(),
                self.port3_combo.get(),
                self.port4_combo.get(),
                self.port5_combo.get(),
                self.port6_combo.get()
            ]

            selected_ports = [port for port in ports if port and port != "None"]

            if not selected_ports:
                messagebox.showerror("Error", "Please select at least one COM port")
                return

            if len(selected_ports) != len(set(selected_ports)):
                messagebox.showerror("Error", "COM ports must be unique")
                return

            baudrate = int(self.baud_combo.get())

            print(f"\nStarting connection to ports: {', '.join(selected_ports)} at {baudrate} baud")

            self.serial_readers = [
                SerialReader(port, baudrate, self.data_queue, self.crc_calculator)
                for port in selected_ports
            ]

            connection_results = [reader.start() for reader in self.serial_readers]

            if not any(connection_results):
                messagebox.showerror("Error", "Failed to connect to any port")
                return

            if not self.send_start_command():
                print("Warning: Failed to send start command to some ports")

            self.is_running = True
            self.paused = False
            self.demo_mode = False
            self.connect_btn.config(text="Disconnect")
            self.demo_btn.state(['disabled'])
            self.pause_btn.state(['!disabled'])
            self.export_btn.state(['!disabled'])

            self.update_plot()

        except Exception as e:
            print(f"Start error: {str(e)}")
            messagebox.showerror("Error", f"Failed to start: {str(e)}")
            self.stop()

    def apply_filter(self, value, channel_idx):
        filter_type = self.filter_combo.get()
        try:
            order = int(self.filter_order_combo.get())
        except:
            order = 1

        ch = self.channels[channel_idx]

        if filter_type == 'LPF':
            try:
                cutoff = float(self.std_cutoff_entry.get())
            except:
                cutoff = 50.0

            if len(ch['timestamps']) > 0:
                time_diff = ch['timestamps'][-1] - ch['timestamps'][-2] if len(ch['timestamps']) > 1 else 0.001
                if time_diff <= 0:
                    time_diff = 0.001
                alpha = 1 / (1 + 1 / (2 * np.pi * cutoff * time_diff))
            else:
                alpha = 1 / (1 + 1 / (2 * np.pi * cutoff * 0.001))

            if ch['lpf_states'] is None or len(ch['lpf_states']) != order:
                ch['lpf_states'] = [0.0] * order

            current_value = value
            for j in range(order):
                state = ch['lpf_states'][j]
                filtered_val = alpha * current_value + (1 - alpha) * state
                ch['lpf_states'][j] = filtered_val
                current_value = filtered_val

            return filtered_val

        elif filter_type == 'HPF':
            try:
                cutoff = float(self.std_cutoff_entry.get())
            except:
                cutoff = 50.0

            if len(ch['timestamps']) > 0:
                time_diff = ch['timestamps'][-1] - ch['timestamps'][-2] if len(ch['timestamps']) > 1 else 0.001
                if time_diff <= 0:
                    time_diff = 0.001
                alpha = 1 / (1 + 1 / (2 * np.pi * cutoff * time_diff))
            else:
                alpha = 1 / (1 + 1 / (2 * np.pi * cutoff * 0.001))

            if ch['hpf_states'] is None or len(ch['hpf_states']) != order:
                ch['hpf_states'] = [{'state': 0.0, 'prev_input': 0.0} for _ in range(order)]

            current_value = value
            for j in range(order):
                state_dict = ch['hpf_states'][j]
                hp = alpha * (state_dict['state'] + current_value - state_dict['prev_input'])
                state_dict['state'] = hp
                state_dict['prev_input'] = current_value
                current_value = hp

            return hp

        elif filter_type == 'BPF':
            try:
                cutoff_low = float(self.bpf_low_entry.get())
                cutoff_high = float(self.bpf_high_entry.get())
            except:
                cutoff_low = 20.0
                cutoff_high = 100.0

            if cutoff_low > cutoff_high:
                cutoff_low, cutoff_high = cutoff_high, cutoff_low

            if len(ch['timestamps']) > 0:
                time_diff = ch['timestamps'][-1] - ch['timestamps'][-2] if len(ch['timestamps']) > 1 else 0.001
                if time_diff <= 0:
                    time_diff = 0.001
                alpha_low = 1 / (1 + 1 / (2 * np.pi * cutoff_low * time_diff))
                alpha_high = 1 / (1 + 1 / (2 * np.pi * cutoff_high * time_diff))
            else:
                alpha_low = 1 / (1 + 1 / (2 * np.pi * cutoff_low * 0.001))
                alpha_high = 1 / (1 + 1 / (2 * np.pi * cutoff_high * 0.001))

            if ch['bpf_sections'] is None or len(ch['bpf_sections']) != order:
                ch['bpf_sections'] = [{'lpf': 0.0, 'hpf': 0.0, 'prev_lpf': 0.0} for _ in range(order)]

            current_value = value
            for j in range(order):
                section = ch['bpf_sections'][j]

                lpf_val = alpha_high * current_value + (1 - alpha_high) * section['lpf']
                section['lpf'] = lpf_val

                hpf_val = alpha_low * (section['hpf'] + lpf_val - section['prev_lpf'])
                section['hpf'] = hpf_val
                section['prev_lpf'] = lpf_val

                current_value = hpf_val

            return hpf_val

        else:
            return value

    def update_plot(self):
        if self.is_running and not self.paused:
            try:
                max_points = 200
                processed = 0

                while not self.data_queue.empty() and processed < max_points:
                    data = self.data_queue.get_nowait()

                    port = data['port']
                    channel_idx = data['channel'] - 1
                    timestamp = data['timestamp']
                    value = data['value']

                    if 0 <= channel_idx < self.num_channels:
                        ch = self.channels[channel_idx]

                        if 'port' not in ch or ch['port'] is None:
                            ch['port'] = port
                            print(f"Channel {channel_idx + 1} assigned to port {port}")

                        if ch['port'] == port:
                            ch['raw_data'].append(value)
                            ch['timestamps'].append(timestamp)
                            ch['filtered_data'].append(self.apply_filter(value, channel_idx))
                            processed += 1

                points_to_show = int(self.points_entry.get())

                for i, (ax, line, ch) in enumerate(zip(self.axes, self.lines, self.channels)):
                    if len(ch['timestamps']) == 0:
                        continue

                    time_array = np.array(ch['timestamps'])
                    data_array = np.array(ch['filtered_data'])

                    start_idx = max(0, len(time_array) - points_to_show)
                    time_axis = time_array[start_idx:]
                    data_axis = data_array[start_idx:]

                    line.set_data(time_axis, data_axis)

                    if len(time_axis) > 1:
                        ax.relim()
                        ax.autoscale_view()
                        ax.set_xlim(time_axis[0], time_axis[-1])

                        y_min, y_max = np.min(data_axis), np.max(data_axis)
                        y_margin = max(0.1 * (y_max - y_min), 0.5)
                        ax.set_ylim(y_min - y_margin, y_max + y_margin)

                self.canvas_time.draw()

            except Exception as e:
                print(f"Plot update error: {str(e)}")

        self.after_id = self.master.after(20, self.update_plot)

    def increase_scale(self, event=None):
        for ax in self.axes:
            y_center = (ax.get_ylim()[0] + ax.get_ylim()[1]) / 2
            y_half_range = (ax.get_ylim()[1] - ax.get_ylim()[0]) * 0.4
            ax.set_ylim(y_center - y_half_range, y_center + y_half_range)
        self.canvas_time.draw()

    def decrease_scale(self, event=None):
        for ax in self.axes:
            y_center = (ax.get_ylim()[0] + ax.get_ylim()[1]) / 2
            y_half_range = (ax.get_ylim()[1] - ax.get_ylim()[0]) * 0.625
            ax.set_ylim(y_center - y_half_range, y_center + y_half_range)
        self.canvas_time.draw()

    def export_data(self):
        try:
            patient_info = {
                "Фамилия": self.surname_entry.get(),
                "Имя": self.name_entry.get(),
                "Отчество": self.patronymic_entry.get(),
                "Пол": self.gender_combo.get(),
                "Год рождения": self.birth_year_entry.get(),
                "Диагноз": self.diagnosis_entry.get(),
                "Дата записи": time.strftime("%Y-%m-%d %H:%M:%S"),
                "Режим": "Демо-режим" if self.demo_mode else "Реальные данные"
            }

            try:
                points_to_save = int(self.points_entry.get())
                if points_to_save <= 0:
                    raise ValueError
            except ValueError:
                messagebox.showerror("Error", "Invalid points value! Please enter a positive integer.")
                return

            points_available = min(points_to_save,
                                   min(len(ch['timestamps']) for ch in self.channels if len(ch['timestamps']) > 0))

            if points_available == 0:
                messagebox.showwarning("No Data", "There is no data to export.")
                return

            data_dict = {}
            for i, ch in enumerate(self.channels):
                if len(ch['timestamps']) > 0:
                    data_dict[f'Ch{i + 1}_Time'] = list(ch['timestamps'])[-points_available:]
                    data_dict[f'Ch{i + 1}_Raw'] = list(ch['raw_data'])[-points_available:]
                    data_dict[f'Ch{i + 1}_Filtered'] = list(ch['filtered_data'])[-points_available:]

            df_data = pd.DataFrame(data_dict)
            df_info = pd.DataFrame(list(patient_info.items()), columns=['Parameter', 'Value'])

            file_path = filedialog.asksaveasfilename(
                defaultextension='.xlsx',
                filetypes=[('Excel files', '*.xlsx'), ('CSV files', '*.csv')],
                title="Save data as"
            )

            if not file_path:
                return

            if file_path.endswith('.csv'):
                df_combined = pd.concat(
                    [df_info, pd.DataFrame([[''] * len(df_info.columns)], columns=df_info.columns), df_data],
                    ignore_index=True)
                df_combined.to_csv(file_path, index=False, encoding='utf-8')
                messagebox.showinfo("Success", f"Data exported successfully to:\n{file_path}")
            else:
                try:
                    with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                        df_info.to_excel(writer,
                                         sheet_name='Patient Info',
                                         index=False,
                                         header=['Parameter', 'Value'])

                        df_data.to_excel(writer,
                                         sheet_name='Measurement Data',
                                         index=False)

                        from openpyxl.utils import get_column_letter

                        worksheet_info = writer.sheets['Patient Info']
                        for idx in range(len(df_info.columns)):
                            col_letter = get_column_letter(idx + 1)
                            max_len = max(
                                df_info.iloc[:, idx].astype(str).apply(len).max(),
                                len(str(df_info.columns[idx]))
                            )
                            worksheet_info.column_dimensions[col_letter].width = max_len + 2

                        worksheet_data = writer.sheets['Measurement Data']
                        for idx in range(len(df_data.columns)):
                            col_letter = get_column_letter(idx + 1)
                            max_len = max(
                                df_data.iloc[:, idx].astype(str).apply(len).max(),
                                len(str(df_data.columns[idx]))
                            )
                            worksheet_data.column_dimensions[col_letter].width = max_len + 2

                    messagebox.showinfo("Success", f"Data exported successfully to:\n{file_path}")

                except Exception as e:
                    messagebox.showerror("Export Error", f"Failed to export data to Excel:\n{str(e)}")

        except Exception as e:
            messagebox.showerror("Export Error", f"Failed to export data:\n{str(e)}")

    def stop(self):
        print("\nStopping application...")

        if self.demo_mode:
            self.stop_demo()

        if not self.demo_mode:
            self.send_stop_command()

        if self.after_id:
            try:
                self.master.after_cancel(self.after_id)
            except:
                pass
            self.after_id = None

        if hasattr(self, 'serial_readers'):
            for reader in self.serial_readers:
                try:
                    reader.stop()
                except Exception as e:
                    print(f"Error stopping reader: {str(e)}")
            self.serial_readers = []

        self.is_running = False
        self.demo_mode = False
        self.connect_btn.config(text="Connect")
        self.demo_btn.config(text="▶ Start Demo")
        self.demo_btn.state(['!disabled'])
        self.pause_btn.state(['disabled'])
        self.export_btn.state(['disabled'])

        print("Application stopped successfully")


if __name__ == "__main__":
    root = tk.Tk()
    root.title("ECG Oscilloscope with Real-Time Analysis - 10 Leads")
    root.geometry("1500x800")
    app = Oscilloscope(root)
    root.mainloop()