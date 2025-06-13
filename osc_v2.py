import matplotlib
matplotlib.use('TkAgg')
from matplotlib.backend_bases import MouseEvent
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import serial
from serial.tools import list_ports
import numpy as np
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from collections import deque
import time
import pandas as pd
import queue
import threading


class Oscilloscope:
    def __init__(self, master):
        self.after_id = None  # Добавлен идентификатор задачи обновления
        self.master = master
        self.ser = None
        self.is_running = False
        self.paused = False
        self.buffer_size = 5000
        self.total_points = 0
        self.num_channels = 4
        self.channels = [
            {
                'raw_data': deque(maxlen=self.buffer_size),
                'filtered_data': deque(maxlen=self.buffer_size),
                'filter_state': 0.0,
                'prev_raw': 0.0,
                'bpf_lpf_state': 0.0,  # Новые состояния для БПФ
                'bpf_hpf_state': 0.0,
                'bpf_prev_lpf': 0.0,
                'color': ['#1f77b4', '#ff7f0e', '#2ca02c', '#d62728'][i]
            } for i in range(self.num_channels)
        ]
        self.timestamps = deque(maxlen=self.buffer_size)
        self.start_time = None
        self.sample_rate = 533.0
        self.sample_period = 1.0 / self.sample_rate
        self.data_queue = queue.Queue()
        self.filter_params = {
            'type': 'None',
            'cutoff': 50.0
        }

        self.setup_gui()
        self.setup_plots()
        self.refresh_ports()

        self.master.bind('+', self.increase_scale)
        self.master.bind('-', self.decrease_scale)

        master.protocol("WM_DELETE_WINDOW", self.on_close)  # Добавляем обработчик закрытия

    def on_close(self):
        """Обработчик закрытия главного окна"""
        self.stop()  # Корректно останавливаем все процессы
        self.master.destroy()  # Уничтожаем окно

    def setup_gui(self):
        main_panel = ttk.PanedWindow(self.master, orient=tk.HORIZONTAL)
        main_panel.pack(fill=tk.BOTH, expand=True)

        # Левая панель с информацией о пациенте
        left_info_panel = ttk.Frame(main_panel, width=250)
        main_panel.add(left_info_panel)

        # Правая панель с графиками и управлением
        right_panel = ttk.Frame(main_panel)
        main_panel.add(right_panel)

        # Информация о пациенте
        patient_frame = ttk.LabelFrame(left_info_panel, text="Информация о пациенте")
        patient_frame.pack(padx=10, pady=10, fill=tk.X)

        ttk.Label(patient_frame, text="Фамилия:").grid(row=0, column=0, sticky=tk.W)
        self.surname_entry = ttk.Entry(patient_frame)
        self.surname_entry.grid(row=0, column=1, sticky=tk.EW, padx=5, pady=2)

        ttk.Label(patient_frame, text="Имя:").grid(row=1, column=0, sticky=tk.W)
        self.name_entry = ttk.Entry(patient_frame)
        self.name_entry.grid(row=1, column=1, sticky=tk.EW, padx=5, pady=2)

        ttk.Label(patient_frame, text="Отчество:").grid(row=2, column=0, sticky=tk.W)
        self.patronymic_entry = ttk.Entry(patient_frame)
        self.patronymic_entry.grid(row=2, column=1, sticky=tk.EW, padx=5, pady=2)

        ttk.Label(patient_frame, text="Пол:").grid(row=3, column=0, sticky=tk.W)
        self.gender_combo = ttk.Combobox(patient_frame, values=["М", "Ж"], width=3)
        self.gender_combo.grid(row=3, column=1, sticky=tk.W, padx=5, pady=2)

        ttk.Label(patient_frame, text="Год рождения:").grid(row=4, column=0, sticky=tk.W)
        self.birth_year_entry = ttk.Entry(patient_frame, width=8)
        self.birth_year_entry.grid(row=4, column=1, sticky=tk.W, padx=5, pady=2)

        ttk.Label(patient_frame, text="Диагноз:").grid(row=5, column=0, sticky=tk.W)
        self.diagnosis_entry = ttk.Entry(patient_frame)
        self.diagnosis_entry.grid(row=5, column=1, sticky=tk.EW, padx=5, pady=2)

        # Переносим остальные элементы управления в правую панель
        self.notebook = ttk.Notebook(right_panel)
        self.notebook.pack(fill=tk.BOTH, expand=True)

        # Time Domain Tab
        time_frame = ttk.Frame(self.notebook)
        self.notebook.add(time_frame, text='Oscilloscope')

        # Control Panel
        control_frame = ttk.Frame(time_frame)
        control_frame.pack(side=tk.TOP, fill=tk.X, padx=5, pady=5)

        left_panel = ttk.Frame(control_frame)
        left_panel.pack(side=tk.LEFT, fill=tk.X, expand=True)

        right_panel = ttk.Frame(control_frame)
        right_panel.pack(side=tk.RIGHT)

        # Port settings
        ttk.Label(left_panel, text="Port:").pack(side=tk.LEFT)
        self.port_combo = ttk.Combobox(left_panel, width=15)
        self.port_combo.pack(side=tk.LEFT)

        ttk.Label(left_panel, text="Baud:").pack(side=tk.LEFT, padx=(10, 0))
        self.baud_combo = ttk.Combobox(left_panel, values=[9600, 19200, 38400, 57600, 115200,250000], width=10)
        self.baud_combo.current(5)
        self.baud_combo.pack(side=tk.LEFT)

        self.connect_btn = ttk.Button(left_panel, text="Connect", command=self.toggle_connection)
        self.connect_btn.pack(side=tk.LEFT, padx=10)

        # Sample settings
        ttk.Label(left_panel, text="Rate (Hz):").pack(side=tk.LEFT)
        self.rate_entry = ttk.Entry(left_panel, width=8)
        self.rate_entry.insert(0, "533")
        self.rate_entry.pack(side=tk.LEFT)
        self.rate_entry.bind("<FocusOut>", self.update_sample_rate)

        ttk.Label(left_panel, text="Points:").pack(side=tk.LEFT)
        self.points_entry = ttk.Entry(left_panel, width=8)
        self.points_entry.insert(0, "1000")
        self.points_entry.pack(side=tk.LEFT)

        # Filter settings
        ttk.Label(left_panel, text="Filter:").pack(side=tk.LEFT)
        self.filter_combo = ttk.Combobox(left_panel, values=['None', 'LPF', 'HPF', 'BPF'], width=6)
        self.filter_combo.current(0)
        self.filter_combo.pack(side=tk.LEFT)
        self.filter_combo.bind("<<ComboboxSelected>>", self.update_filter_ui)

        # Фрейм для стандартного cutoff (LPF/HPF)
        self.std_cutoff_frame = ttk.Frame(left_panel)
        self.std_cutoff_label = ttk.Label(self.std_cutoff_frame, text="Cutoff (Hz):")
        self.std_cutoff_label.pack(side=tk.LEFT)
        self.std_cutoff_entry = ttk.Entry(self.std_cutoff_frame, width=8)
        self.std_cutoff_entry.insert(0, "50")
        self.std_cutoff_entry.pack(side=tk.LEFT)
        self.std_cutoff_frame.pack(side=tk.LEFT)

        # Фрейм для BPF cutoff (нижний и верхний)
        self.bpf_cutoff_frame = ttk.Frame(left_panel)
        self.bpf_low_label = ttk.Label(self.bpf_cutoff_frame, text="Low:")
        self.bpf_low_label.pack(side=tk.LEFT)
        self.bpf_low_entry = ttk.Entry(self.bpf_cutoff_frame, width=6)
        self.bpf_low_entry.insert(0, "20")
        self.bpf_low_entry.pack(side=tk.LEFT)
        self.bpf_high_label = ttk.Label(self.bpf_cutoff_frame, text="High:")
        self.bpf_high_label.pack(side=tk.LEFT)
        self.bpf_high_entry = ttk.Entry(self.bpf_cutoff_frame, width=6)
        self.bpf_high_entry.insert(0, "100")
        self.bpf_high_entry.pack(side=tk.LEFT)
        self.bpf_cutoff_frame.pack_forget()

        # Buttons
        self.export_btn = ttk.Button(right_panel, text="📤 Export", command=self.export_data)
        self.export_btn.pack(side=tk.RIGHT, padx=5)
        self.export_btn.state(['disabled'])

        self.pause_btn = ttk.Button(right_panel, text="⏸ Stop", command=self.toggle_pause)
        self.pause_btn.pack(side=tk.RIGHT, padx=5)
        self.pause_btn.state(['disabled'])

    def setup_plots(self):
        time_frame = self.notebook.winfo_children()[0]
        self.fig_time = plt.figure(figsize=(10, 8), dpi=100)
        self.axes = []
        self.lines = []

        for i in range(self.num_channels):
            ax = self.fig_time.add_subplot(self.num_channels, 1, i + 1)
            line, = ax.plot([], [], lw=1, color=self.channels[i]['color'])
            ax.set_ylabel(f'Ch {i + 1}')
            ax.grid(True)
            self.axes.append(ax)
            self.lines.append(line)

        self.annotations = []
        for i, ax in enumerate(self.axes):
            # Создаем скрытую аннотацию
            annotation = ax.annotate('',
                                     xy=(0, 0),
                                     xytext=(5, -15 if i < self.num_channels - 1 else 5),  # Позиция надписи
                                     textcoords='offset points',
                                     bbox=dict(boxstyle="round", fc="w", alpha=0.8),
                                     arrowprops=dict(arrowstyle="->")
                                     )
            annotation.set_visible(False)
            self.annotations.append(annotation)

        # Связываем событие движения мыши
        self.fig_time.canvas.mpl_connect('motion_notify_event', self.on_mouse_move)
        self.fig_time.canvas.mpl_connect('figure_leave_event', self.on_leave_figure)

        self.axes[-1].set_xlabel('Time (seconds)')
        self.canvas_time = FigureCanvasTkAgg(self.fig_time, master=time_frame)
        self.canvas_time.get_tk_widget().pack(fill=tk.BOTH, expand=True)

    def on_mouse_move(self, event):
        """Обработчик движения мыши над графиком"""
        if event.inaxes is None:
            for ann in self.annotations:
                ann.set_visible(False)
            self.canvas_time.draw_idle()
            return

            # Проверяем, над каким графиком мышь
        for i, ax in enumerate(self.axes):
            if ax == event.inaxes:
                # Получаем данные для текущего графика
                line = self.lines[i]
                xdata = line.get_xdata()
                ydata = line.get_ydata()

                if len(xdata) == 0:
                    continue

                # Находим ближайшую точку данных
                idx = np.abs(xdata - event.xdata).argmin()
                x = xdata[idx]
                y = ydata[idx]

                # Обновляем аннотацию
                self.annotations[i].xy = (x, y)
                self.annotations[i].set_text(f'x={x:.2f}s, y={y:.2f}mV')
                self.annotations[i].set_visible(True)
            else:
                # Скрываем аннотации для других графиков
                self.annotations[i].set_visible(False)

        self.canvas_time.draw_idle()

    def on_leave_figure(self, event):
        """Скрываем аннотации при выходе за пределы графика"""
        for ann in self.annotations:
            ann.set_visible(False)
        self.canvas_time.draw_idle()

    # Добавим метод update_filter_ui для управления видимостью элементов:
    def update_filter_ui(self, event=None):
        filter_type = self.filter_combo.get()
        if filter_type == 'BPF':
            self.std_cutoff_frame.pack_forget()
            self.bpf_cutoff_frame.pack(side=tk.LEFT)
        else:
            self.bpf_cutoff_frame.pack_forget()
            self.std_cutoff_frame.pack(side=tk.LEFT)
    def update_sample_rate(self, event=None):
        try:
            new_rate = float(self.rate_entry.get())
            if new_rate <= 0:
                raise ValueError
            self.sample_rate = new_rate
            self.sample_period = 1.0 / self.sample_rate
        except:
            messagebox.showerror("Error", "Invalid sample rate value!")
            self.rate_entry.delete(0, tk.END)
            self.rate_entry.insert(0, str(self.sample_rate))

    def refresh_ports(self):
        ports = [port.device for port in list_ports.comports()]
        self.port_combo['values'] = ports
        if ports:
            self.port_combo.current(0)

    def toggle_connection(self):
        if self.is_running:
            self.stop()
        else:
            self.start()

    def toggle_pause(self):
        if self.paused:
            # Возобновление измерений с полным сбросом данных
            self._reset_measurement()
            if self.ser and self.ser.is_open:
                self.ser.reset_input_buffer()  # Очистка буфера COM-порта
            self.paused = False
            self.pause_btn.config(text="⏸ Stop")
            self.start_time = time.time()  # Новый отсчёт времени
        else:
            # Пауза без очистки данных
            self.paused = True
            self.pause_time = time.time()
            self.pause_btn.config(text="▶ Start")

    def _reset_measurement(self):
        """Сброс всех данных и графиков, как при новом подключении"""
        self.total_points = 0
        self.timestamps.clear()
        self.data_queue.queue.clear()

        # Очистка данных каналов и сброс фильтров
        for ch in self.channels:
            ch['raw_data'].clear()
            ch['filtered_data'].clear()
            ch['filter_state'] = 0.0
            ch['prev_raw'] = 0.0
            ch['bpf_lpf_state'] = 0.0
            ch['bpf_hpf_state'] = 0.0
            ch['bpf_prev_lpf'] = 0.0

        # Очистка графиков
        for line in self.lines:
            line.set_data([], [])
        self.canvas_time.draw()

    def start(self):
        try:
            # Полная очистка предыдущих данных
            self.stop()  # Убедимся, что предыдущее соединение закрыто
            self.total_points = 0
            self.timestamps.clear()
            for ch in self.channels:
                ch['raw_data'].clear()
                ch['filtered_data'].clear()
                ch['filter_state'] = 0.0  # Сброс состояния фильтра
                ch['prev_raw'] = 0.0

            # Инициализация соединения
            self.ser = serial.Serial(
                self.port_combo.get(),
                int(self.baud_combo.get()),
                timeout=0.1
            )
            self.ser.reset_input_buffer()  # Очистка буфера COM-порта
            self.is_running = True
            self.paused = False  # Сброс состояния паузы
            self.pause_btn.state(['!disabled'])
            self.export_btn.state(['!disabled'])
            self.connect_btn.config(text="Disconnect")
            self.start_time = time.time()  # Новый отсчет времени

            for line in self.lines:
                line.set_data([], [])
            self.canvas_time.draw()

            # Запуск потока
            self.read_thread = threading.Thread(target=self.read_from_port, daemon=True)
            self.read_thread.start()
            self.update_plot()
        except Exception as e:
            print("Error:", e)
            messagebox.showerror("Connection Error", str(e))

    def read_from_port(self):
        while self.is_running:
            if self.ser and self.ser.in_waiting:
                try:
                    raw_data = self.ser.readline()
                    line = raw_data.decode('ascii', errors='ignore').strip()
                    if line:
                        print(f"Получено: {line}")  # Логируем сырые данные
                        parts = line.split(';')
                        if len(parts) == 4:
                            try:
                                values = list(map(float, parts))
                                self.data_queue.put(values)
                            except ValueError:
                                print(f"Ошибка в данных: {line}")
                        else:
                            print(f"Некорректная строка: {line}")
                except Exception as e:
                    print("Ошибка:", e)

    def apply_filter(self, values):
        filter_type = self.filter_combo.get()
        filtered = []

        for i, value in enumerate(values):
            ch = self.channels[i]
            if filter_type == 'LPF':
                try:
                    cutoff = float(self.std_cutoff_entry.get())
                except:
                    cutoff = 50.0
                alpha = 1 / (1 + 1 / (2 * np.pi * cutoff / self.sample_rate))
                ch['filter_state'] = alpha * value + (1 - alpha) * ch['filter_state']
                filtered.append(ch['filter_state'])
            elif filter_type == 'HPF':
                try:
                    cutoff = float(self.std_cutoff_entry.get())
                except:
                    cutoff = 50.0
                alpha = 1 / (1 + 1 / (2 * np.pi * cutoff / self.sample_rate))
                hp = alpha * (ch['filter_state'] + value - ch['prev_raw'])
                ch['filter_state'] = hp
                ch['prev_raw'] = value
                filtered.append(hp)
            elif filter_type == 'BPF':
                try:
                    cutoff_low = float(self.bpf_low_entry.get())
                    cutoff_high = float(self.bpf_high_entry.get())
                except:
                    cutoff_low = 20.0
                    cutoff_high = 100.0
                if cutoff_low > cutoff_high:
                    cutoff_low, cutoff_high = cutoff_high, cutoff_low

                # Применяем LPF с верхней частотой
                alpha_high = 1 / (1 + 1 / (2 * np.pi * cutoff_high / self.sample_rate))
                lpf_value = alpha_high * value + (1 - alpha_high) * ch['bpf_lpf_state']
                ch['bpf_lpf_state'] = lpf_value

                # Применяем HPF с нижней частотой
                alpha_low = 1 / (1 + 1 / (2 * np.pi * cutoff_low / self.sample_rate))
                hpf_value = alpha_low * (ch['bpf_hpf_state'] + lpf_value - ch['bpf_prev_lpf'])
                ch['bpf_hpf_state'] = hpf_value
                ch['bpf_prev_lpf'] = lpf_value
                filtered.append(hpf_value)
            else:
                filtered.append(value)
        return filtered



    def update_plot(self):
        if self.is_running and not self.paused:
            try:
                max_points = 200
                processed = 0

                while not self.data_queue.empty() and processed < max_points:
                    raw_values = self.data_queue.get_nowait()
                    current_time = self.total_points * self.sample_period

                    self.total_points += 1

                    filtered_values = self.apply_filter(raw_values)

                    self.timestamps.append(current_time)
                    for i in range(self.num_channels):
                        self.channels[i]['raw_data'].append(raw_values[i])
                        self.channels[i]['filtered_data'].append(filtered_values[i])
                    processed += 1

                points_to_show = min(int(self.points_entry.get()), len(self.timestamps))
                if points_to_show > 0:
                    time_axis = np.array(self.timestamps)[-points_to_show:]

                    # Обновляем данные для всех графиков
                    for i, ax in enumerate(self.axes):
                        data_axis = np.array(self.channels[i]['filtered_data'])[-points_to_show:]
                        self.lines[i].set_data(time_axis, data_axis)

                        # Автомасштабирование Y для каждого канала
                        ax.relim()
                        ax.autoscale_view(scalex=False, scaley=True)

                    # Синхронизация осей X для всех subplots
                    if len(time_axis) > 1:
                        x_min = max(0, time_axis[0])
                        x_max = time_axis[-1] + 0.1 * (time_axis[-1] - time_axis[0])
                        if x_min == x_max:
                            x_min -= 0.1
                            x_max += 0.1
                    elif len(time_axis) == 1:
                        x_min = time_axis[0] - 0.5
                        x_max = time_axis[0] + 0.5
                    else:
                        x_min, x_max = 0, 1  # Значения по умолчанию

                    # Устанавливаем одинаковые пределы для всех осей X
                    for ax in self.axes:
                        ax.set_xlim(x_min, x_max)

                    self.canvas_time.draw()


            except Exception as e:
                print(f"Update error: {e}")



        # Сохраняем идентификатор задачи
        self.after_id = self.master.after(10, self.update_plot)

    def increase_scale(self, event=None):
        """Увеличение масштаба осей Y с сохранением центра"""
        for ax in self.axes:
            y_center = (ax.get_ylim()[0] + ax.get_ylim()[1]) / 2  # Текущий центр
            y_half_range = (ax.get_ylim()[1] - ax.get_ylim()[0]) * 0.4  # 80% от текущего диапазона
            ax.set_ylim(y_center - y_half_range, y_center + y_half_range)
        self.canvas_time.draw()

    def decrease_scale(self, event=None):
        """Уменьшение масштаба осей Y с сохранением центра"""
        for ax in self.axes:
            y_center = (ax.get_ylim()[0] + ax.get_ylim()[1]) / 2  # Текущий центр
            y_half_range = (ax.get_ylim()[1] - ax.get_ylim()[0]) * 0.625  # 125% от текущего диапазона
            ax.set_ylim(y_center - y_half_range, y_center + y_half_range)
        self.canvas_time.draw()

    def export_data(self):
        try:
            # Собираем информацию о пациенте и параметрах
            patient_info = {
                "Фамилия": self.surname_entry.get(),
                "Имя": self.name_entry.get(),
                "Отчество": self.patronymic_entry.get(),
                "Пол": self.gender_combo.get(),
                "Год рождения": self.birth_year_entry.get(),
                "Диагноз": self.diagnosis_entry.get(),
                "Частота дискретизации (Гц)": self.sample_rate,
                "Дата записи": time.strftime("%Y-%m-%d %H:%M:%S")
            }

            # Получаем данные для экспорта
            try:
                points_to_save = int(self.points_entry.get())
                if points_to_save <= 0:
                    raise ValueError
            except ValueError:
                messagebox.showerror("Error", "Invalid points value! Please enter a positive integer.")
                return

            if not self.timestamps:
                messagebox.showwarning("No Data", "There is no data to export.")
                return

            points_available = min(points_to_save, len(self.timestamps))
            if points_available == 0:
                messagebox.showwarning("No Data", "There is no data to export.")
                return

            # Формируем данные измерений
            data_dict = {'Time (s)': list(self.timestamps)[-points_available:]}
            for i in range(self.num_channels):
                data_dict[f'Ch{i + 1}_Raw'] = list(self.channels[i]['raw_data'])[-points_available:]
                data_dict[f'Ch{i + 1}_Filtered'] = list(self.channels[i]['filtered_data'])[-points_available:]

            df_data = pd.DataFrame(data_dict)
            df_info = pd.DataFrame(list(patient_info.items()), columns=['Parameter', 'Value'])

            # Сохраняем файл
            # Сохраняем файл
            file_path = filedialog.asksaveasfilename(
                defaultextension='.xlsx',
                filetypes=[('Excel files', '*.xlsx'), ('CSV files', '*.csv')],
                title="Save data as"
            )

            if not file_path:
                return

            if file_path.endswith('.csv'):
                # ... [код для CSV] ...
                messagebox.showinfo("Success", f"Data exported successfully to:\n{file_path}")
            else:
                try:
                    # Используем openpyxl как движок
                    with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                        # Лист с информацией о пациенте
                        df_info.to_excel(writer,
                                         sheet_name='Patient Info',
                                         index=False,
                                         header=['Parameter', 'Value'])

                        # Лист с данными измерений
                        df_data.to_excel(writer,
                                         sheet_name='Measurement Data',
                                         index=False)

                        # Настраиваем ширину колонок
                        from openpyxl.utils import get_column_letter

                        # Для листа Patient Info
                        worksheet_info = writer.sheets['Patient Info']
                        for idx in range(len(df_info.columns)):
                            col_letter = get_column_letter(idx + 1)
                            # Определяем максимальную длину в колонке
                            max_len = max(
                                df_info.iloc[:, idx].astype(str).apply(len).max(),
                                len(str(df_info.columns[idx]))
                            )
                            worksheet_info.column_dimensions[col_letter].width = max_len + 2

                        # Для листа Measurement Data
                        worksheet_data = writer.sheets['Measurement Data']
                        for idx in range(len(df_data.columns)):
                            col_letter = get_column_letter(idx + 1)
                            max_len = max(
                                df_data.iloc[:, idx].astype(str).apply(len).max(),
                                len(str(df_data.columns[idx]))
                            )
                            worksheet_data.column_dimensions[col_letter].width = max_len + 2

                    messagebox.showinfo("Success", f"Data exported successfully to:\n{file_path}")

                except Exception as e:
                    messagebox.showerror("Export Error", f"Failed to export data to Excel:\n{str(e)}")

        except Exception as e:
            messagebox.showerror("Export Error", f"Failed to export data:\n{str(e)}")

    def stop(self):
        """Полная остановка всех процессов и освобождение ресурсов"""
        if self.is_running or self.ser is not None:
            self.is_running = False

            # 1. Отмена запланированных задач обновления
            if self.after_id:
                self.master.after_cancel(self.after_id)
                self.after_id = None

            # 2. Закрытие COM-порта с обработкой ошибок
            if self.ser and self.ser.is_open:
                try:
                    self.ser.close()
                    print("COM port closed successfully")
                except serial.SerialException as e:
                    print(f"Error closing port: {str(e)}")
                finally:
                    self.ser = None

            # 3. Остановка потока чтения данных
            if hasattr(self, 'read_thread'):
                try:
                    self.read_thread.join(timeout=0.5)
                    if self.read_thread.is_alive():
                        print("Warning: Read thread not terminated properly")
                except Exception as e:
                    print(f"Thread join error: {str(e)}")

            # 4. Обновление состояния интерфейса
            self.connect_btn.config(text="Connect")
            self.pause_btn.state(['disabled'])
            self.export_btn.state(['disabled'])

        # 5. Очистка буферов данных
        self.data_queue.queue.clear()
        self.timestamps.clear()
        for ch in self.channels:
            ch['raw_data'].clear()
            ch['filtered_data'].clear()


if __name__ == "__main__":
    root = tk.Tk()
    root.title("4-Channel Oscilloscope")
    root.geometry("1200x800")
    app = Oscilloscope(root)
    root.mainloop()